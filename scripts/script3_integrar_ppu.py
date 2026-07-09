from __future__ import annotations

import logging
import re
import sys as _sys
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, simpledialog

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


if getattr(_sys, "frozen", False):
    BASE_DIR = Path(_sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parents[1]

INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
PPU_FILE = INPUT_DIR / "Premissa_Receita_OLK.xlsx"
CRONOGRAMA_FILE = INPUT_DIR / "cronograma_outlook_validacao.xlsx"


def setup_logging() -> logging.Logger:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log_path = OUTPUT_DIR / f"script3_integrar_ppu_{datetime.now():%Y%m%d_%H%M%S}.log"

    logger = logging.getLogger("output_outlook.script3")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    sh = logging.StreamHandler()
    sh.setFormatter(formatter)
    logger.addHandler(sh)

    logger.info("Log inicializado em: %s", log_path)
    return logger


def normalize_name(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = text.replace("a", "a")
    text = text.replace("á", "a").replace("à", "a").replace("â", "a").replace("ã", "a")
    text = text.replace("é", "e").replace("ê", "e")
    text = text.replace("í", "i")
    text = text.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    text = text.replace("ú", "u")
    text = text.replace("ç", "c")
    return " ".join(text.split())


def find_column(df: pd.DataFrame, aliases: list[str], required: bool = True) -> str | None:
    normalized = {normalize_name(c): str(c) for c in df.columns}
    for alias in aliases:
        col = normalized.get(normalize_name(alias))
        if col:
            return col

    if required:
        raise KeyError(f"Nao foi possivel localizar coluna. Tentativas: {aliases}")
    return None


def parse_yyyymm_from_filename(path: Path) -> str:
    stem = path.stem
    digits = "".join(ch for ch in stem if ch.isdigit())
    if len(digits) >= 6:
        return digits[-6:]
    now = datetime.now()
    return f"{now.year}{now.month:02d}"


def ask_cronograma_file() -> Path | None:
    root = tk.Tk()
    root.withdraw()

    initial = OUTPUT_DIR / "cronograma_outlook_validacao_AAAAMM.xlsx"
    path_str = simpledialog.askstring(
        "Arquivo cronograma",
        "Informe o caminho completo do arquivo do Script 1 (.xlsx):",
        initialvalue=str(initial),
        parent=root,
    )
    root.destroy()

    if not path_str:
        return None

    path = Path(path_str.strip().strip('"'))
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")
    return path


def map_tipo_ppu_to_cronograma(value: object) -> str:
    text = normalize_name(value)
    mapping = {
        "uep": "UEP",
        "producao": "UEP",
        "producao offshore": "UEP",
        "sonda": "SONDA",
        "perfuracao": "SONDA",
        "perf": "SONDA",
        "ums": "UMS",
    }
    return mapping.get(text, str(value).strip().upper() if value is not None else "")


def to_numeric_series(series: pd.Series) -> pd.Series:
    def _normalize_numeric_text(value: object) -> str:
        text = "" if value is None else str(value).strip()
        if not text:
            return ""

        text = text.replace(" ", "")
        has_comma = "," in text
        has_dot = "." in text

        # Suporta formatos mistos: 1.234,56 / 1,234.56 / 7074,804 / 7074.804
        if has_comma and has_dot:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "")
                text = text.replace(",", ".")
            else:
                text = text.replace(",", "")
        elif has_comma:
            text = text.replace(",", ".")

        return text

    cleaned = series.map(_normalize_numeric_text)
    return pd.to_numeric(cleaned, errors="coerce")


def load_ppu(logger: logging.Logger) -> pd.DataFrame:
    if not PPU_FILE.exists():
        raise FileNotFoundError(
            f"Arquivo PPU nao encontrado em {PPU_FILE}. Coloque o arquivo Premissa_Receita_OLK.xlsx na pasta input/."
        )

    df = pd.read_excel(PPU_FILE)
    col_capping = find_column(df, ["Capping?", "Capping"], required=False)
    if col_capping:
        before = len(df)
        df = df[df[col_capping].map(normalize_name) == "nao"].copy()
        logger.info("Filtro Capping aplicado (%s=Nao): %d -> %d linhas", col_capping, before, len(df))

    logger.info("Base PPU carregada: %s (linhas=%d)", PPU_FILE, len(df))
    return df


def build_integrated(df_cronograma: pd.DataFrame, df_ppu: pd.DataFrame, logger: logging.Logger) -> pd.DataFrame:
    col_pep_crono = find_column(df_cronograma, ["PEP raiz", "PEP"])
    col_tipo_crono = find_column(df_cronograma, ["TIPO", "Tipo"])
    col_atendimento_crono = find_column(df_cronograma, ["Atendimento"], required=False)
    col_competencia_crono = find_column(df_cronograma, ["MesReferencia", "Competência", "Competencia"], required=False)
    col_dias = find_column(df_cronograma, ["DiasOperadosMes", "Dias"])
    col_prefixo = find_column(df_cronograma, ["Prefixo SAP", "Prefixo"], required=False)

    col_pep_ppu = find_column(df_ppu, ["PEP raiz", "PEP"])
    col_tipo_ppu = find_column(df_ppu, ["Tipo de UM", "Tipo de Unidade", "Tipo Unidade", "Tipo", "Tipo PPU"])
    col_fator_conv = find_column(
        df_ppu,
        [
            "Fator de Conversao",
            "Fator de Conversão",
            "Fator Conversao",
            "Fator de Conversão (quantidade)",
            "Fator",
        ],
        required=False,
    )
    col_preco = find_column(df_ppu, ["Preco Unitario", "Preço Unitário", "Preco", "PU"], required=False)
    col_reajuste = find_column(
        df_ppu,
        ["Fator de Reajuste", "Fator Reajuste", "Fator Reajuste Cliente", "Reajuste"],
        required=False,
    )
    col_demanda = find_column(df_ppu, ["Demanda"], required=False)

    crono = df_cronograma.copy()
    ppu = df_ppu.copy()
    ppu_columns = list(ppu.columns)

    if col_atendimento_crono:
        before = len(crono)
        crono = crono[crono[col_atendimento_crono].map(normalize_name) != "petrobras"].copy()
        logger.info(
            "Filtro Atendimento aplicado (%s != Petrobras): %d -> %d linhas",
            col_atendimento_crono,
            before,
            len(crono),
        )

    crono["_key_pep"] = crono[col_pep_crono].map(normalize_name)
    crono["_key_tipo"] = crono[col_tipo_crono].fillna("").astype(str).str.strip().str.upper()

    ppu["_key_pep"] = ppu[col_pep_ppu].map(normalize_name)
    ppu["_key_tipo"] = ppu[col_tipo_ppu].map(map_tipo_ppu_to_cronograma)

    # Mantem como base as colunas da Premissa_Receita_OLK e traz do cronograma
    # apenas os campos necessários para os cálculos finais.
    crono_merge_cols = ["_key_pep", "_key_tipo"]
    if col_competencia_crono:
        crono_merge_cols.append(col_competencia_crono)
    crono_merge_cols.append(col_dias)
    if col_prefixo:
        crono_merge_cols.append(col_prefixo)

    merged = ppu.merge(
        crono[crono_merge_cols],
        how="left",
        on=["_key_pep", "_key_tipo"],
        suffixes=("", "_CRONO"),
    )

    dias = to_numeric_series(merged[col_dias])
    fator_conv = to_numeric_series(merged[col_fator_conv]) if col_fator_conv else pd.Series(0, index=merged.index)
    preco = to_numeric_series(merged[col_preco]) if col_preco else pd.Series(0, index=merged.index)
    reajuste = to_numeric_series(merged[col_reajuste]) if col_reajuste else pd.Series(0, index=merged.index)

    dias = dias.fillna(0)
    fator_conv = fator_conv.fillna(0)
    preco = preco.fillna(0)
    reajuste = reajuste.fillna(0)

    if col_competencia_crono:
        merged["Competência"] = merged[col_competencia_crono]
        before = len(merged)
        merged = merged[
            merged["Competência"].notna()
            & (merged["Competência"].astype(str).str.strip() != "")
        ].copy()
        logger.info("Filtro Competencia nao vazia aplicado: %d -> %d linhas", before, len(merged))
    merged["Dias"] = dias
    merged["Quant. Medida"] = fator_conv * dias
    merged["Montante em Moeda Original"] = preco * reajuste * merged["Quant. Medida"]

    if col_prefixo:
        merged["Unidade Marítima"] = merged[col_prefixo]
    else:
        merged["Unidade Marítima"] = ""

    # Saída final: colunas da Premissa + apenas os 5 campos acordados
    output_cols = list(ppu_columns)
    for extra_col in ["Competência", "Dias", "Unidade Marítima", "Quant. Medida", "Montante em Moeda Original"]:
        if extra_col not in output_cols:
            output_cols.append(extra_col)

    merged = merged[output_cols]
    logger.info("Integracao cronograma + PPU concluida (linhas=%d)", len(merged))
    return merged


def _apply_table_medium8(ws, table_name: str) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return

    if ws.tables:
        for existing_name in list(ws.tables.keys()):
            del ws.tables[existing_name]

    table_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium8",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _ensure_unique_headers(ws) -> None:
    seen: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col_idx).value
        base = "" if raw is None else str(raw).strip()
        if not base:
            base = f"Coluna_{col_idx}"

        count = seen.get(base, 0)
        if count == 0:
            new_name = base
        else:
            new_name = f"{base}_{count + 1}"

        seen[base] = count + 1
        ws.cell(row=1, column=col_idx, value=new_name)


def apply_output_format(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["cronograma_ppu"]

    max_row = ws.max_row
    max_col = ws.max_column

    _ensure_unique_headers(ws)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill_group2 = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_fill_group3 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    ws.freeze_panes = "A2"

    header_map = {str(c.value): c.column for c in ws[1] if c.value is not None}

    group2_cols = ["Competência", "Dias", "Unidade Marítima"]
    group3_cols = ["Quant. Medida", "Montante em Moeda Original"]

    for col_name in group2_cols:
        col_idx = header_map.get(col_name)
        if not col_idx:
            continue
        ws.cell(row=1, column=col_idx).fill = header_fill_group2

    for col_name in group3_cols:
        col_idx = header_map.get(col_name)
        if not col_idx:
            continue
        ws.cell(row=1, column=col_idx).fill = header_fill_group3

    competencia_col = header_map.get("Competência")
    if competencia_col:
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=competencia_col, max_col=competencia_col):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "DD/MM/YYYY"

    for col_name, fmt in [
        ("Dias", "#,##0.00"),
        ("Quant. Medida", "#,##0.00"),
        ("Montante em Moeda Original", "#,##0.00"),
    ]:
        col_idx = header_map.get(col_name)
        if col_idx:
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = fmt

    validar_col = header_map.get("Validar")
    if validar_col and max_row >= 2:
        validar_letter = ws.cell(row=1, column=validar_col).column_letter
        validar_range = f"{validar_letter}2:{validar_letter}{max_row}"
        ws.conditional_formatting.add(
            validar_range,
            FormulaRule(formula=[f"=${validar_letter}2=TRUE"], fill=green_fill),
        )
        ws.conditional_formatting.add(
            validar_range,
            FormulaRule(formula=[f"=${validar_letter}2=FALSE"], fill=red_fill),
        )

    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    _apply_table_medium8(ws, "TabelaCronogramaPPU")

    wb.save(workbook_path)
    wb.close()


def main() -> None:
    logger = setup_logging()

    try:
        cronograma_path = CRONOGRAMA_FILE
        if not cronograma_path.exists():
            raise FileNotFoundError(
                f"Arquivo de entrada nao encontrado: {cronograma_path}. "
                "Execute o Script 1 e garanta que ele salve em input/cronograma_outlook_validacao.xlsx."
            )

        logger.info("Arquivo de entrada: %s", cronograma_path)
        df_cronograma = pd.read_excel(cronograma_path, sheet_name="cronograma_base")

        df_ppu = load_ppu(logger)
        df_integrado = build_integrated(df_cronograma, df_ppu, logger)

        yyyymm = parse_yyyymm_from_filename(cronograma_path)
        output_path = OUTPUT_DIR / f"cronograma_ppu_integrado_{yyyymm}.xlsx"

        try:
            with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
                df_integrado.to_excel(writer, sheet_name="cronograma_ppu", index=False)
        except PermissionError:
            output_path = OUTPUT_DIR / f"cronograma_ppu_integrado_{yyyymm}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            logger.warning("Arquivo de saida estava aberto. Usando fallback: %s", output_path)
            with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
                df_integrado.to_excel(writer, sheet_name="cronograma_ppu", index=False)

        apply_output_format(output_path)
        logger.info("Arquivo gerado: %s", output_path)

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Script 3", f"Arquivo gerado com sucesso:\n{output_path}", parent=root)
        root.destroy()
    except Exception as exc:
        logger.exception("Falha na execucao do Script 3: %s", exc)
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Script 3 - Erro", str(exc), parent=root)
            root.destroy()
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()
