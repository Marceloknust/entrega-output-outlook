"""
================================================================================
SCRIPT 1 - GERAR ARQUIVO DE VALIDAÇÃO
================================================================================

OBJETIVO:
--------
Gerar arquivo "cronograma_outlook_validacao_AAAAMM.xlsx" para validação do time
Comercial, contendo 4 abas: Base Dados, cronograma_base, TD e ajuste.

ETAPAS:
-------
1. Usuário seleciona ANO e MÊS (interface gráfica)
2. Lê arquivo "Anexo I - Base de Dados Demanda Primária mês AAAA.xlsx" buscando o arquivo mais recente na pasta input/
3. Aplica filtro: TIPO = "SONDA", "UEP" ou "UMS"
4. Expande para cronograma base (replicando expandir_cronoweb.py)
5. Cria chave de busca:
    - Prioridade: Prefixo SAP + Bloco Ajustado + TIPO
    - Fallback: Bacia Ajustada + Bloco Ajustado + TIPO
6. Aplica DE_PARA usando a nova chave
7. Gera arquivo com 4 abas:
   - "Base Dados": cópia integral do original
   - "cronograma_base": 14 colunas (sem colunas de ajuste)
   - "TD": tabela dinâmica com slicers
   - "ajuste": dados únicos com colunas amarelas para preenchimento
8. Envia e-mail para time Comercial (com caixa de diálogo)

FORMATAÇÕES ESPECÍFICAS:
-----------------------
- Aba "ajuste": colunas amarelas (#FFFF00) para preenchimento
- Aba "cronograma_base": formatação padrão
- Aba "TD": tabela dinâmica com slicers
- Aba "Base Dados": formatação original mantida

ARQUIVO DE SAÍDA:
----------------
cronograma_outlook_validacao_AAAAMM.xlsx (onde AAAAMM = ano e mês selecionados)

================================================================================
DATA CRIAÇÃO: Junho/2026
ÚLTIMA ATUALIZAÇÃO: Junho/2026
================================================================================
"""
#importando as bibliotecas necessarias
from __future__ import annotations

import argparse
import calendar
import getpass
import logging
import shutil
import time
from dataclasses import dataclass
import re
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk

import pandas as pd
import win32com.client  # type: ignore[import]
from openpyxl.formatting.rule import FormulaRule
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

## Configuracoes de diretorios
import sys as _sys
if getattr(_sys, "frozen", False):
    # Executavel PyInstaller: caminhos relativos ao .exe
    BASE_DIR = Path(_sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parents[1]
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"

## Configuracoes de arquivos
BASE_DADOS_FILE = INPUT_DIR / "Base_Dados.xlsx"
DE_PARA_FILE = INPUT_DIR / "DE_PARA.xlsx"

## Configuracoes de e-mail
DEFAULT_TO = [
    "marcelo.knust.prestserv@transpetro.com.br",
    #"barcanias@transpetro.com.br",
    #"williamarthur@transpetro.com.br",
    #"mrnobrega@transpetro.com.br",
]
DEFAULT_CC = [
    #"marcelo.knust.prestserv@transpetro.com.br",
    "felipedan@transpetro.com.br",
]

## Configuracoes de log definindo colunas utilizadas em cronograma_base e ajuste
CRONOGRAMA_COLS = [
    "TIPO",
    "Exclusões",
    "Bacia Ajustada",
    "Bloco Ajustado",
    "Prefixo SAP",
    "Nome da Locação",
    "Nome do Poço",
    "MesReferencia",
    "InicioNoMes",
    "FimNoMes",
    "DiasOperadosMes",
    "DiasOperadosMes_Rateado",
    # "DiasIntervaloTotal",  # removido temporariamente
    "Atendimento",
    "Contrato",
    "Modalidade",
    "PEP raiz",
]

AJUSTE_COLS = [
    "TIPO",
    "Exclusões",
    "Bacia Ajustada",
    "Bloco Ajustado",
    "Prefixo SAP",
    "Nome da Locação",
    "Nome do Poço",
    "PEP raiz",
    "Validação",
    "Atendimento_ajustado",
    "Contrato_ajustado",
    "Modalidade_ajustado",
    "Inicio_ajustado",
    "Final_ajustado",
]

## Configuracoes de periodo
@dataclass(frozen=True)
class MonthSlice:
    mes_referencia: date
    inicio_no_mes: date
    fim_no_mes: date
    dias_operados_mes: int

##Gerando o logger para registrar as informacoes do script
def setup_logging() -> logging.Logger:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log_path = OUTPUT_DIR / f"script1_gerar_validacao_{datetime.now():%Y%m%d_%H%M%S}.log"

    logger = logging.getLogger("output_outlook.script1")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    logger.info("Log inicializado em: %s", log_path)
    return logger

## Configuracoes de normalizacao de nomes
def normalize_name(value: object) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = text.replace("á", "a").replace("à", "a").replace("â", "a").replace("ã", "a")
    text = text.replace("é", "e").replace("ê", "e")
    text = text.replace("í", "i")
    text = text.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    text = text.replace("ú", "u")
    text = text.replace("ç", "c")
    return " ".join(text.split())

### Funcoes de manipulacao de datas
def excel_value_to_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None

## Funções de manipulacao de datas para calculo de periodos customizados
def first_day_of_month(value: date) -> date:
    return value.replace(day=1)

def last_day_of_month(value: date) -> date:
    return value.replace(day=calendar.monthrange(value.year, value.month)[1])

def next_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)

def previous_month(value: date) -> date:
    if value.month == 1:
        return date(value.year - 1, 12, 1)
    return date(value.year, value.month - 1, 1)

def month_ref_start_for_date(value: date) -> date:
    if value.day >= 26:
        return next_month(first_day_of_month(value))
    return first_day_of_month(value)

def build_month_period(month_ref_start: date) -> tuple[date, date, date]:
    prev_month = previous_month(month_ref_start)
    period_start = prev_month.replace(day=26)
    period_end = month_ref_start.replace(day=25)
    month_ref_end = last_day_of_month(month_ref_start)
    return period_start, period_end, month_ref_end

def split_interval_by_custom_month(start_date: date, end_date: date) -> list[MonthSlice]:
    slices: list[MonthSlice] = []
    cursor = month_ref_start_for_date(start_date)
    end_cursor = month_ref_start_for_date(end_date)

    while cursor <= end_cursor:
        period_start, period_end, month_ref_end = build_month_period(cursor)
        slice_start = max(start_date, period_start)
        slice_end = min(end_date, period_end)

        if slice_start <= slice_end:
            slices.append(
                MonthSlice(
                    mes_referencia=month_ref_end,
                    inicio_no_mes=slice_start,
                    fim_no_mes=slice_end,
                    dias_operados_mes=(slice_end - slice_start).days + 1,
                )
            )

        cursor = next_month(cursor)

    return slices

## Funcoes de interface grafica para selecao de ano e mes
def pick_year_month(timeout_seconds: int = 600) -> tuple[int, int, int, int] | None:
    current = date.today().year
    years = [str(y) for y in range(current - 5, current + 6)]
    months = [
        ("01", "Janeiro"),
        ("02", "Fevereiro"),
        ("03", "Marco"),
        ("04", "Abril"),
        ("05", "Maio"),
        ("06", "Junho"),
        ("07", "Julho"),
        ("08", "Agosto"),
        ("09", "Setembro"),
        ("10", "Outubro"),
        ("11", "Novembro"),
        ("12", "Dezembro"),
    ]

    root = tk.Tk()
    root.title("Selecao de Periodo - output_outlook")
    root.geometry("460x420")
    root.resizable(False, False)
    root.attributes("-topmost", True)
    root.lift()
    root.focus_force()

    ttk.Label(root, text="Ano:").pack(anchor="w", padx=16, pady=(16, 4))
    selected_year = tk.StringVar(value=str(current))
    ttk.Combobox(root, values=years, textvariable=selected_year, state="readonly").pack(fill=tk.BOTH, padx=16)

    ttk.Label(root, text="Mes Inicial:").pack(anchor="w", padx=16, pady=(12, 4))
    month_display = [f"{num} - {name}" for num, name in months]
    selected_month = tk.StringVar(value=month_display[date.today().month - 1])
    ttk.Combobox(root, values=month_display, textvariable=selected_month, state="readonly").pack(fill=tk.BOTH, padx=16)

    ttk.Label(root, text="Ano Final:").pack(anchor="w", padx=16, pady=(12, 4))
    selected_year_end = tk.StringVar(value=str(current))
    ttk.Combobox(root, values=years, textvariable=selected_year_end, state="readonly").pack(fill=tk.BOTH, padx=16)

    ttk.Label(root, text="Mes Final:").pack(anchor="w", padx=16, pady=(12, 4))
    selected_month_end = tk.StringVar(value="12 - Dezembro")
    ttk.Combobox(root, values=month_display, textvariable=selected_month_end, state="readonly").pack(fill=tk.BOTH, padx=16)

    result: dict[str, int] = {}
    remaining_seconds = max(30, int(timeout_seconds))
    timer_var = tk.StringVar()

    def format_remaining(seconds: int) -> str:
        mins = seconds // 60
        secs = seconds % 60
        return f"Tempo para selecao: {mins:02d}:{secs:02d}"

    def tick() -> None:
        nonlocal remaining_seconds
        timer_var.set(format_remaining(remaining_seconds))
        if remaining_seconds <= 0:
            messagebox.showwarning("Tempo esgotado", "Tempo para selecao encerrado. Execute novamente para escolher o periodo.")
            root.destroy()
            return
        remaining_seconds -= 1
        root.after(1000, tick)

    def confirm() -> None:
        try:
            result["year"] = int(selected_year.get())
            result["month"] = int(selected_month.get().split(" - ", 1)[0])
            result["year_end"] = int(selected_year_end.get())
            result["month_end"] = int(selected_month_end.get().split(" - ", 1)[0])

            start_ref = date(result["year"], result["month"], 1)
            end_ref = date(result["year_end"], result["month_end"], 1)
            if end_ref < start_ref:
                raise ValueError("Periodo final anterior ao inicial.")
            root.destroy()
        except Exception:
            messagebox.showerror("Erro", "Selecione ano/mes inicial e final validos.")

    def cancel() -> None:
        root.destroy()

    def extend_time() -> None:
        nonlocal remaining_seconds
        remaining_seconds += 300
        timer_var.set(format_remaining(remaining_seconds))

    buttons = ttk.Frame(root)
    ttk.Label(root, textvariable=timer_var).pack(anchor="w", padx=16, pady=(8, 2))
    ttk.Button(root, text="+5 min", command=extend_time).pack(anchor="w", padx=16, pady=(0, 6))
    buttons.pack(fill=tk.BOTH, padx=16, pady=(4, 14))
    ttk.Button(buttons, text="Cancelar", command=cancel).pack(side=tk.RIGHT)
    ttk.Button(buttons, text="Confirmar", command=confirm).pack(side=tk.RIGHT, padx=(0, 8))

    tick()
    root.mainloop()

    if "year" not in result or "month" not in result or "year_end" not in result or "month_end" not in result:
        return None
    return result["year"], result["month"], result["year_end"], result["month_end"]

### Funcoes de manipulacao de arquivos e envio de e-mail
def _extract_year_month_from_filename(path: Path) -> tuple[int, int] | None:
    name = normalize_name(path.stem)
    months = {
        "janeiro": 1,
        "fevereiro": 2,
        "marco": 3,
        "abril": 4,
        "maio": 5,
        "junho": 6,
        "julho": 7,
        "agosto": 8,
        "setembro": 9,
        "outubro": 10,
        "novembro": 11,
        "dezembro": 12,
    }
    month_pattern = "|".join(months.keys())

    # Ex.: "Abril 2026"
    m = re.search(rf"({month_pattern})\D+(20\d{{2}})", name)
    if m:
        return int(m.group(2)), months[m.group(1)]

    # Ex.: "2026 Abril"
    m = re.search(rf"(20\d{{2}})\D+({month_pattern})", name)
    if m:
        return int(m.group(1)), months[m.group(2)]

    # Ex.: "202604"
    m = re.search(r"(20\d{2})(0[1-9]|1[0-2])", name)
    if m:
        return int(m.group(1)), int(m.group(2))

    # Ex.: "04-2026" ou "04_2026"
    m = re.search(r"(0[1-9]|1[0-2])\D+(20\d{2})", name)
    if m:
        return int(m.group(2)), int(m.group(1))

    return None

## Funcoes de manipulacao de arquivos e envio de e-mail
def resolve_base_dados_file() -> Path:
    candidates = [
        p for p in INPUT_DIR.glob("*.xlsx") if normalize_name(p.name) != normalize_name(DE_PARA_FILE.name)
    ]
    if not candidates:
        raise FileNotFoundError(
            f"Arquivo de base nao encontrado em {INPUT_DIR}. Esperado: {BASE_DADOS_FILE.name}"
        )

    # Busca arquivos no formato esperado de nomenclatura da base.
    base_like = [p for p in candidates if "base" in normalize_name(p.stem) and "dados" in normalize_name(p.stem)]
    pool = base_like if base_like else candidates

    ranked: list[tuple[tuple[int, int], Path]] = []
    for path in pool:
        year_month = _extract_year_month_from_filename(path)
        if year_month is not None:
            ranked.append((year_month, path))

    if ranked:
        # Maior (ano, mes) = nomenclatura mais nova.
        return max(ranked, key=lambda item: item[0])[1]

    # Fallback para cenários sem mês/ano no nome.
    return sorted(pool, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def read_excel_with_fallback(path: Path, preferred_sheet: str) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True)
    sheet = preferred_sheet if preferred_sheet in wb.sheetnames else wb.sheetnames[0]
    wb.close()
    return pd.read_excel(path, sheet_name=sheet)


def find_column(df: pd.DataFrame, aliases: list[str], required: bool = True) -> str | None:
    normalized = {normalize_name(c): str(c) for c in df.columns}
    for alias in aliases:
        col = normalized.get(normalize_name(alias))
        if col:
            return col

    if required:
        raise KeyError(f"Nao foi possivel localizar coluna. Tentativas: {aliases}")
    return None


def month_name_pt(month: int) -> str:
    names = [
        "Janeiro",
        "Fevereiro",
        "Marco",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    return names[month - 1]


def _clean_key_part(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _lookup_key_preferred(prefixo_unidade: object, bloco: object, tipo: object | None = None) -> tuple[str, str, str, str] | None:
    p = _clean_key_part(prefixo_unidade)
    b = _clean_key_part(bloco)
    t = _clean_key_part(tipo)
    if p and b and t:
        return ("prefixo", p, b, t.upper())
    return None


def _lookup_key_fallback(bacia: object, bloco: object, tipo: object | None = None) -> tuple[str, str, str, str] | None:
    ba = _clean_key_part(bacia)
    bl = _clean_key_part(bloco)
    t = _clean_key_part(tipo)
    if ba and bl and t:
        return ("bacia", ba, bl, t.upper())
    return None


def _expand_de_para_tipos(row: pd.Series, col_sonda: str | None, col_uep: str | None, col_ums: str | None) -> list[str]:
    tipo_map = [
        ("SONDA", col_sonda),
        ("UEP", col_uep),
        ("UMS", col_ums),
    ]
    selected: list[str] = []
    for tipo, col_name in tipo_map:
        if col_name is None:
            continue
        if normalize_name(row.get(col_name)) == "sim":
            selected.append(tipo)

    # Compatibilidade com o arquivo atual: se nao houver colunas por tipo,
    # aplica a linha para os tres tipos.
    if not selected:
        selected = ["SONDA", "UEP", "UMS"]
    return selected


def _normalize_integral(value: object) -> str:
    text = normalize_name(value)
    if text == "sim":
        return "sim"
    if text == "nao":
        return "nao"
    return ""


def build_de_para_map(
    df_de_para: pd.DataFrame,
    logger: logging.Logger,
) -> tuple[dict[tuple[str, str, str, str], dict[str, object]], set[tuple[str, str]]]:
    col_bacia = find_column(df_de_para, ["Bacia Ajustado", "Bacia Ajustada"])
    col_bloco = find_column(df_de_para, ["Bloco Ajustado"])
    col_unidade = find_column(df_de_para, ["Unidade Marítima", "Unidade Maritima"], required=False)
    col_integral = find_column(df_de_para, ["Integral?", "Integral"])
    col_atendimento = find_column(df_de_para, ["Atendimento"], required=False)
    col_contrato = find_column(df_de_para, ["Contrato"], required=False)
    col_modalidade = find_column(df_de_para, ["Modalidade"], required=False)
    col_pep = find_column(df_de_para, ["PEP raiz", "PEP"], required=False)
    col_sonda = find_column(df_de_para, ["Sonda", "SONDA"], required=False)
    col_uep = find_column(df_de_para, ["UEP"], required=False)
    col_ums = find_column(df_de_para, ["UMS"], required=False)

    result: dict[tuple[str, str, str, str], dict[str, object]] = {}
    sim_bacia_bloco: set[tuple[str, str]] = set()
    nao_bacia_bloco: set[tuple[str, str]] = set()
    duplicates: dict[tuple[str, str, str, str], list[int]] = {}
    invalid_integral_rows: list[int] = []

    for idx, row in df_de_para.iterrows():
        row_number = idx + 2
        unidade_val = row.get(col_unidade) if col_unidade else None
        bloco_val = row.get(col_bloco)
        bacia_val = row.get(col_bacia)
        integral_val = _normalize_integral(row.get(col_integral))

        if not integral_val:
            invalid_integral_rows.append(row_number)
            logger.error(
                "DE_PARA linha %d com Integral? inválido/vazio. Valor recebido: %r",
                row_number,
                row.get(col_integral),
            )
            continue

        payload = {
            "Atendimento": row.get(col_atendimento) if col_atendimento else None,
            "Contrato": row.get(col_contrato) if col_contrato else None,
            "Modalidade": row.get(col_modalidade) if col_modalidade else None,
            "PEP raiz": row.get(col_pep) if col_pep else None,
        }

        tipos_expandidos = _expand_de_para_tipos(row, col_sonda, col_uep, col_ums)
        base_kind = "bacia" if integral_val == "sim" else "prefixo"
        base_value = bacia_val if integral_val == "sim" else unidade_val

        if integral_val == "sim":
            key_desc = "Bacia Ajustada + Bloco Ajustado + TIPO"
            bacia_clean = _clean_key_part(bacia_val)
            bloco_clean = _clean_key_part(bloco_val)
            if bacia_clean and bloco_clean:
                sim_bacia_bloco.add((bacia_clean, bloco_clean))
        else:
            key_desc = "Unidade Maritima + Bloco Ajustado + TIPO"
            bacia_clean = _clean_key_part(bacia_val)
            bloco_clean = _clean_key_part(bloco_val)
            if bacia_clean and bloco_clean:
                nao_bacia_bloco.add((bacia_clean, bloco_clean))

        key_base = _clean_key_part(base_value)
        if not key_base:
            logger.error(
                "DE_PARA linha %d sem chave base completa para Integral=%s (%s).",
                row_number,
                integral_val,
                key_desc,
            )
            continue

        for tipo in tipos_expandidos:
            key = (base_kind, key_base, _clean_key_part(bloco_val), tipo)
            if key in result:
                duplicates.setdefault(key, []).append(row_number)
                continue
            result[key] = payload

    if invalid_integral_rows:
        logger.error(
            "DE_PARA com %d linhas inválidas em Integral? (ver erros acima).",
            len(invalid_integral_rows),
        )

    if duplicates:
        details = []
        for dup_key, rows in list(duplicates.items())[:10]:
            details.append(f"{dup_key} linhas_duplicadas={rows}")
        logger.warning(
            "DE_PARA possui %d chaves duplicadas para a regra de Integral?. "
            "Mantendo sempre a primeira ocorrência. Exemplos: %s",
            len(duplicates),
            " | ".join(details),
        )

    # Bloqueia fallback por bacia+bloco somente quando nao houver linha Integral=Sim para o par.
    nao_only_bacia_bloco = nao_bacia_bloco - sim_bacia_bloco
    return result, nao_only_bacia_bloco


def build_cronograma_base(
    df_base: pd.DataFrame,
    de_para_map: dict[tuple[str, str, str, str], dict[str, object]],
    nao_scoped_bacia_bloco: set[tuple[str, str]],
    ano_base: int,
    mes_base: int,
    ano_final: int,
    mes_final: int,
    logger: logging.Logger | None = None,
) -> pd.DataFrame:
    col_tipo = find_column(df_base, ["TIPO", "Tipo"])
    col_exclusoes = find_column(df_base, ["Exclusões", "Exclusoes", "Q"], required=False)
    col_bacia = find_column(df_base, ["Bacia Ajustada", "Bacia Ajustado"])
    col_bloco = find_column(df_base, ["Bloco Ajustado"])
    col_prefixo = find_column(df_base, ["Prefixo SAP", "Prefixo"])
    col_nome_locacao = find_column(df_base, ["Nome da Locação", "Nome da Locacao"])
    col_nome_poco = find_column(df_base, ["Nome do Poço", "Nome do Poco"])
    col_inicio = find_column(df_base, ["Data inicio", "Data Inicio", "Inicio", "Inicio Programado", "Data inicial"])
    col_fim = find_column(df_base, ["Data termino", "Data fim", "Fim", "Fim Programado", "Data final"])

    # AnoMesBase define o corte inicial; processa do proprio mes inicial ate o periodo final informado.
    first_ref_month_start = date(ano_base, mes_base, 1)
    min_ref = last_day_of_month(first_ref_month_start)
    max_ref = last_day_of_month(date(ano_final, mes_final, 1))

    rows_out: list[dict[str, object]] = []
    no_match_keys: set[tuple[str, str, str]] = set()
    for _, row in df_base.iterrows():
        tipo_val = _clean_key_part(row.get(col_tipo)).upper()
        if tipo_val not in {"SONDA", "UEP", "UMS"}:
            continue

        # Regra: manter apenas registros com Exclusões vazio.
        exclusao_val = _clean_key_part(row.get(col_exclusoes, "")) if col_exclusoes else ""
        if exclusao_val:
            continue

        start_date = excel_value_to_date(row.get(col_inicio))
        end_date = excel_value_to_date(row.get(col_fim))
        if start_date is None or end_date is None or end_date < start_date:
            continue

        month_slices = split_interval_by_custom_month(start_date, end_date)
        # total_days = sum(item.dias_operados_mes for item in month_slices)  # DiasIntervaloTotal desativado

        for item in month_slices:
            if item.mes_referencia < min_ref or item.mes_referencia > max_ref:
                continue

            bacia_val = _clean_key_part(row.get(col_bacia, ""))
            bloco_val = _clean_key_part(row.get(col_bloco, ""))
            prefixo_val = _clean_key_part(row.get(col_prefixo, ""))

            preferred_key = _lookup_key_preferred(prefixo_val, bloco_val, tipo_val)
            fallback_key = _lookup_key_fallback(bacia_val, bloco_val, tipo_val)

            # Prioridade: Prefixo SAP + Bloco + TIPO.
            if preferred_key is not None and preferred_key in de_para_map:
                lookup_key = preferred_key
            elif (bacia_val, bloco_val) in nao_scoped_bacia_bloco:
                # Para pares Bacia+Bloco em regra Integral=Não, nao pode usar fallback por Bacia+Bloco.
                lookup_key = None
            elif fallback_key is not None and fallback_key in de_para_map:
                lookup_key = fallback_key
            else:
                lookup_key = None

            d = de_para_map.get(lookup_key, {}) if lookup_key else {}
            has_match = lookup_key is not None

            if not has_match:
                no_match_keys.add((bacia_val, bloco_val, prefixo_val))

            atendimento_val = _clean_key_part(d.get("Atendimento"))
            contrato_val = _clean_key_part(d.get("Contrato"))
            modalidade_val = _clean_key_part(d.get("Modalidade"))
            pep_val = _clean_key_part(d.get("PEP raiz"))

            if atendimento_val == "" and contrato_val == "" and modalidade_val == "":
                atendimento_val = "petrobras"
                contrato_val = "petrobras"
                modalidade_val = "petrobras"

            contrato_norm = normalize_name(contrato_val)
            if (
                tipo_val == "UMS"
                and "controle de emergencia" in contrato_norm
                and "buzios" not in contrato_norm
                and "forno" not in contrato_norm
            ):
                atendimento_val = "Petrobras"

            rows_out.append(
                {
                    "TIPO": row.get(col_tipo),
                    "Exclusões": row.get(col_exclusoes) if col_exclusoes else None,
                    "Bacia Ajustada": bacia_val,
                    "Bloco Ajustado": bloco_val,
                    "Prefixo SAP": row.get(col_prefixo),
                    "Nome da Locação": row.get(col_nome_locacao),
                    "Nome do Poço": row.get(col_nome_poco),
                    "MesReferencia": item.mes_referencia,
                    "InicioNoMes": item.inicio_no_mes,
                    "FimNoMes": item.fim_no_mes,
                    "DiasOperadosMes": item.dias_operados_mes,
                    "DiasOperadosMes_Rateado": float(item.dias_operados_mes),
                    # "DiasIntervaloTotal": total_days,  # desativado temporariamente
                    "Atendimento": atendimento_val,
                    "Contrato": contrato_val,
                    "Modalidade": modalidade_val,
                    "PEP raiz": pep_val,
                }
            )

    if not rows_out:
        raise ValueError("Nenhuma linha encontrada para o período selecionado.")

    if no_match_keys and logger:
        logger.warning(
            "%d combinacoes sem match no DE_PARA (Bacia - Bloco - Prefixo):",
            len(no_match_keys),
        )
        for bacia, bloco, prefixo in sorted(no_match_keys):
            logger.warning("  Registro sem match no DE_PARA: %s - %s - %s", bacia, bloco, prefixo)

    return pd.DataFrame(rows_out)[CRONOGRAMA_COLS]


def build_ajuste_sheet(df_cronograma: pd.DataFrame) -> pd.DataFrame:
    keys = ["Bacia Ajustada", "Bloco Ajustado", "Prefixo SAP"]
    sem_match = (
        df_cronograma["Atendimento"].fillna("").astype(str).str.strip().eq("")
        & df_cronograma["Contrato"].fillna("").astype(str).str.strip().eq("")
        & df_cronograma["Modalidade"].fillna("").astype(str).str.strip().eq("")
    )
    status = (
        df_cronograma.assign(_sem_match=sem_match)
        .groupby(keys, as_index=False)["_sem_match"]
        .max()
    )
    periodos = (
        df_cronograma.groupby(keys, as_index=False)
        .agg(
            {
                "InicioNoMes": "min",
                "FimNoMes": "max",
            }
        )
        .rename(
            columns={
                "InicioNoMes": "Inicio_ajustado",
                "FimNoMes": "Final_ajustado",
            }
        )
    )

    ajuste = df_cronograma.drop_duplicates(
        subset=keys,
        keep="first",
    ).copy()
    ajuste = ajuste.merge(status, on=keys, how="left")
    ajuste = ajuste.merge(periodos, on=keys, how="left")
    ajuste["Validação"] = ajuste["_sem_match"].map({True: "Não", False: "Sim"}).fillna("Sim")
    ajuste = ajuste.drop(columns=["_sem_match"])

    ajuste["Atendimento_ajustado"] = ajuste.apply(
        lambda r: r["Atendimento"] if r["Validação"] == "Sim" else "",
        axis=1,
    )
    ajuste["Contrato_ajustado"] = ajuste.apply(
        lambda r: r["Contrato"] if r["Validação"] == "Sim" else "",
        axis=1,
    )
    ajuste["Modalidade_ajustado"] = ajuste.apply(
        lambda r: r["Modalidade"] if r["Validação"] == "Sim" else "",
        axis=1,
    )

    campos_ajustados = ["Atendimento_ajustado", "Contrato_ajustado", "Modalidade_ajustado"]
    todos_em_branco = (
        ajuste[campos_ajustados]
        .fillna("")
        .astype(str)
        .apply(lambda col: col.str.strip())
        .eq("")
        .all(axis=1)
    )
    ajuste.loc[todos_em_branco, campos_ajustados] = "Petrobras"

    return ajuste[AJUSTE_COLS]


def write_output_workbook(base_dados_path: Path, df_cronograma: pd.DataFrame, df_ajuste: pd.DataFrame, output_path: Path) -> None:
    wb = load_workbook(base_dados_path)
    base_sheet = wb["Base Dados"] if "Base Dados" in wb.sheetnames else wb[wb.sheetnames[0]]

    for sheet_name in list(wb.sheetnames):
        if wb[sheet_name] is not base_sheet:
            del wb[sheet_name]

    base_sheet.title = "Base Dados"

    ws_crono = wb.create_sheet("cronograma_base")
    ws_crono.append(CRONOGRAMA_COLS)
    for row in df_cronograma.itertuples(index=False, name=None):
        ws_crono.append(list(row))

    wb.create_sheet("TD")
    ws_ajuste = wb.create_sheet("ajuste")
    ws_ajuste.append(AJUSTE_COLS)
    for row in df_ajuste.itertuples(index=False, name=None):
        ws_ajuste.append(list(row))

    wb.save(output_path)
    wb.close()


def apply_cronograma_basic_format(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["cronograma_base"]

    max_row = ws.max_row
    max_col = ws.max_column

    table_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    ws.freeze_panes = "A2"

    header_map = {str(c.value): c.column for c in ws[1] if c.value is not None}
    for col_name in ["MesReferencia", "InicioNoMes", "FimNoMes", "InicioNoMes_ajustado", "FimNoMes_ajustado"]:
        col_idx = header_map.get(col_name)
        if col_idx:
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = "DD/MM/YYYY"

    for col_name in ["DiasOperadosMes", "DiasOperadosMes_Rateado"]:  # "DiasIntervaloTotal" desativado
        col_idx = header_map.get(col_name)
        if col_idx:
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = "#,##0.00"

    validado_col = header_map.get("Validado")
    ajustado_cols = [
        header_map.get("InicioNoMes_ajustado"),
        header_map.get("FimNoMes_ajustado"),
        header_map.get("Atendimento_ajustado"),
        header_map.get("Contrato_ajustado"),
        header_map.get("Modalidade_ajustado"),
    ]

    if validado_col:
        validado_letter = ws.cell(row=1, column=validado_col).column_letter
        validado_range = f"{validado_letter}2:{validado_letter}{max_row}"

        dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=False)
        dv.error = "Use apenas Sim ou Não."
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)
        dv.add(validado_range)

        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=validado_col)
            if cell.value == "Não":
                cell.fill = green_fill
            elif cell.value == "Sim":
                cell.fill = red_fill

        ws.conditional_formatting.add(
            validado_range,
            FormulaRule(formula=[f'${validado_letter}2="Não"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            validado_range,
            FormulaRule(formula=[f'${validado_letter}2="Sim"'], fill=red_fill),
        )

    for col_idx in ajustado_cols:
        if not col_idx:
            continue
        for row_idx in range(2, max_row + 1):
            ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 34)

    _apply_table_medium8(ws, "TabelaCronogramaBase")

    wb.save(workbook_path)
    wb.close()


AJUSTE_FILLABLE_COLS = {
    "Atendimento_ajustado",
    "Contrato_ajustado",
    "Modalidade_ajustado",
    "Inicio_ajustado",
    "Final_ajustado",
}


def _apply_table_medium8(ws, table_name: str) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return

    if ws.tables:
        for existing_name in list(ws.tables.keys()):
            del ws.tables[existing_name]

    table_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium8",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def apply_ajuste_format(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["ajuste"]

    max_row = ws.max_row
    max_col = ws.max_column

    table_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_map = {str(c.value): c.column for c in ws[1] if c.value is not None}
    fillable_cols = {col_idx for name, col_idx in header_map.items() if name in AJUSTE_FILLABLE_COLS}

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            if cell.column in fillable_cols:
                cell.fill = yellow_fill

    validacao_col = header_map.get("Validação")
    if validacao_col:
        validacao_letter = ws.cell(row=1, column=validacao_col).column_letter
        validacao_range = f"{validacao_letter}2:{validacao_letter}{max_row}"

        dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=False)
        dv.error = "Use apenas Sim ou Não."
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)
        dv.add(validacao_range)

        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=validacao_col)
            if cell.value == "Sim":
                cell.fill = green_fill
            elif cell.value == "Não":
                cell.fill = red_fill

        ws.conditional_formatting.add(
            validacao_range,
            FormulaRule(formula=[f'${validacao_letter}2="Sim"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            validacao_range,
            FormulaRule(formula=[f'${validacao_letter}2="Não"'], fill=red_fill),
        )

    ws.freeze_panes = "A2"

    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 34)

    _apply_table_medium8(ws, "TabelaAjuste")

    wb.save(workbook_path)
    wb.close()


def copy_output_to_input(output_path: Path, logger: logging.Logger) -> Path:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    input_copy_path = INPUT_DIR / "cronograma_outlook_validacao.xlsx"
    try:
        shutil.copy2(output_path, input_copy_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Nao foi possivel sobrescrever {input_copy_path}. Feche o arquivo no Excel e execute novamente."
        ) from exc

    logger.info("Copia automatica atualizada para entrada do Script 3: %s", input_copy_path)
    return input_copy_path


def build_td_pivot_with_slicers(workbook_path: Path, logger: logging.Logger) -> None:
    import os
    import subprocess
    import tempfile
    import zipfile

    # Estrategia de persistencia: gera pivot no arquivo aberto e salva uma
    # copia temporaria (equivale ao fluxo manual de salvar com novo nome).
    source_path = workbook_path.resolve()
    temp_save_path = source_path.with_name(f"{source_path.stem}__pivot_tmp{source_path.suffix}")
    file_path_escaped = str(source_path).replace("'", "''")
    saveas_path_escaped = str(temp_save_path).replace("'", "''")

    ps_script = f"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
    $wb = $excel.Workbooks.Open('{file_path_escaped}')
    $wsData = $wb.Worksheets('cronograma_base')
    $wsTD  = $wb.Worksheets('TD')
    $wsTD.Cells.Clear()

    # Remove caches de slicer preexistentes no arquivo para evitar conflito
    for ($ci = $wb.SlicerCaches.Count; $ci -ge 1; $ci--) {{
        try {{ $wb.SlicerCaches.Item($ci).Delete() }} catch {{}}
    }}

    $lastRow = $wsData.UsedRange.Rows.Count
    $lastCol = $wsData.UsedRange.Columns.Count
    $sourceRef = "cronograma_base!R1C1:R$($lastRow)C$($lastCol)"
    Write-Host "SOURCE: $sourceRef"
    $cache = $wb.PivotCaches().Create(1, $sourceRef)
    $pt = $cache.CreatePivotTable($wsTD.Range('A3'), 'PivotCronograma')
    Write-Host "CREATED: $($wsTD.PivotTables().Count)"
    $pt.PivotFields('Bloco Ajustado').Orientation = 1
    $pt.PivotFields('Prefixo SAP').Orientation     = 1
    $pt.PivotFields('Nome da Locação').Orientation  = 1
    $pt.PivotFields('Nome do Poço').Orientation     = 1
    $pfMes = $pt.PivotFields('MesReferencia')
    $pfMes.Orientation = 2
    [void]$pt.AddDataField($pt.PivotFields('DiasOperadosMes'), 'Soma DiasOperadosMes', -4157)
    # Forca formato local para evitar exibicao literal "yyyy" nos rotulos
    try {{ $pfMes.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    try {{ $pfMes.DataRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    try {{ $pfMes.LabelRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    $pt.RefreshTable() | Out-Null
    try {{ $pfMes.DataRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    try {{ $pfMes.LabelRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}

    $pivotRange = $pt.TableRange2
    $anchorLeft = $pivotRange.Left + $pivotRange.Width + 24
    $anchorTop = $pivotRange.Top
    $slicerWidth = 180
    $slicerHeight = 120
    $hGap = 10
    $vGap = 10
    $slicerFields = @('TIPO','Contrato','Modalidade','Atendimento')

    # Definicao clara da grade
    $columns = 2
    $rows = 2

    # Remove slicers antigos da aba TD para evitar sobreposicao entre execucoes
    for ($si = $wsTD.Shapes.Count; $si -ge 1; $si--) {{
        try {{
            $sh = $wsTD.Shapes.Item($si)
            if ($sh.Name -like 'Slicer_*' -or $sh.Name -like 'SL_*') {{
                $sh.Delete()
            }}
        }} catch {{}}
    }}

    # Posicionamento dos slicers
    for ($i = 0; $i -lt $slicerFields.Count; $i++) {{
        $field = $slicerFields[$i]
        try {{
            # Calcula linha e coluna (ordem: esquerda->direita, cima->baixo)
            $row = [math]::Floor($i / $columns)
            $col = $i % $columns

            # Cria o slicer
            $pfSlicer = $pt.PivotFields($field)
            $cacheName = ('SC_' + ($field -replace '[^A-Za-z0-9]', '') + '_' + $i)
            try {{
                $sc = $wb.SlicerCaches.Add2($pt, $pfSlicer, $cacheName)
            }} catch {{
                $sc = $wb.SlicerCaches.Add($pt, $pfSlicer, $cacheName)
            }}

            $sl = $sc.Slicers.Add($wsTD)

            # Posiciona na grade usando ancoras de celula (mais estavel no COM)
            $anchorCell = $wsTD.Cells((2 + ($row * 9)), (10 + ($col * 4)))
            $sl.Left = $anchorCell.Left
            $sl.Top = $anchorCell.Top
            $sl.Width = $slicerWidth
            $sl.Height = $slicerHeight

            Write-Host "Slicer '$field' posicionado em Linha $row, Coluna $col"

        }} catch {{
            Write-Warning "Erro ao criar slicer para '$field': $_"
        }}
    }}
    $pivotCount = $wsTD.PivotTables().Count
    Write-Host "PIVOT_BEFORE_CLOSE: $pivotCount"
    $wb.SaveCopyAs('{saveas_path_escaped}')
    if (-not (Test-Path '{saveas_path_escaped}')) {{
        throw 'SaveCopyAs nao gerou arquivo.'
    }}
    Write-Host "SAVED_PIVOT_COUNT: $pivotCount"
    $wb.Close($false)
}} finally {{
    $excel.Quit()
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""

    ps_file = Path(tempfile.mktemp(suffix=".ps1"))
    ps_file.write_text(ps_script, encoding="utf-8-sig")
    try:
        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(ps_file)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
        stdout = (result.stdout or "").strip()
        stderr = (result.stderr or "").strip()
        logger.info("TD PowerShell stdout: %s", stdout)
        if stderr:
            logger.warning("TD PowerShell stderr: %s", stderr)
        if result.returncode != 0:
            raise RuntimeError(f"PowerShell falhou (rc={result.returncode}): {stderr}")
        if "SAVED_PIVOT_COUNT: 0" in stdout or "SAVED_PIVOT_COUNT:" not in stdout:
            raise RuntimeError(f"Pivot nao foi salva. Output: {stdout}")
        if not temp_save_path.exists():
            raise RuntimeError("PowerShell nao gerou arquivo temporario com pivot.")

        with zipfile.ZipFile(temp_save_path, "r") as zf:
            has_pivot_xml = any(name.startswith("xl/pivotTables/") for name in zf.namelist())
        if not has_pivot_xml:
            raise RuntimeError("Arquivo temporario salvo sem artefatos de pivot (xl/pivotTables).")

        os.replace(temp_save_path, source_path)
        logger.info("Aba TD criada como Tabela Dinâmica com filtros/slicers.")
    finally:
        if temp_save_path.exists():
            try:
                temp_save_path.unlink()
            except Exception:
                pass
        try:
            ps_file.unlink()
        except Exception:
            pass


def _split_email_list(raw_value: str) -> list[str]:
    return [item.strip() for item in raw_value.split(";") if item.strip()]


def _validate_email_list(recipients: list[str]) -> tuple[bool, list[str]]:
    pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    invalid = [email for email in recipients if not pattern.match(email)]
    return len(invalid) == 0, invalid


def ask_and_confirm_recipients(logger: logging.Logger) -> tuple[str, str, bool] | None:
    root = tk.Tk()
    root.withdraw()

    to_default = ";".join(DEFAULT_TO)
    cc_default = ";".join(DEFAULT_CC)

    to_value = simpledialog.askstring("Destino", "Destinatários (Para), separados por ';':", initialvalue=to_default)
    if to_value is None:
        root.destroy()
        logger.info("Coleta de e-mails cancelada pelo usuário.")
        return

    cc_value = simpledialog.askstring("Cópia", "Destinatários em cópia (Cc), separados por ';':", initialvalue=cc_default)
    if cc_value is None:
        root.destroy()
        logger.info("Coleta de e-mails cancelada pelo usuário.")
        return

    to_list = _split_email_list(to_value)
    cc_list = _split_email_list(cc_value)
    all_recipients = to_list + cc_list

    if not to_list:
        messagebox.showerror("E-mails inválidos", "Informe ao menos um destinatário em Para.")
        root.destroy()
        logger.warning("Lista de destinatários em Para vazia.")
        return None

    ok, invalids = _validate_email_list(all_recipients)
    if not ok:
        messagebox.showerror("E-mails inválidos", "Foram encontrados e-mails inválidos:\n" + "\n".join(invalids))
        root.destroy()
        logger.warning("E-mails inválidos informados: %s", invalids)
        return None

    confirm_msg = (
        "Confirma os destinatários abaixo para uso no envio final?\n\n"
        f"Para: {'; '.join(to_list)}\n"
        f"Cc: {'; '.join(cc_list) if cc_list else '(vazio)'}"
    )
    confirmed = messagebox.askyesno("Confirmar destinatários", confirm_msg)
    if not confirmed:
        root.destroy()
        logger.info("Destinatários não confirmados pelo usuário.")
        return None

    should_send_at_end = messagebox.askyesno(
        "Envio no Final",
        "Deseja enviar o e-mail no final do processo?",
    )

    root.destroy()
    logger.info("Destinatários validados e confirmados para uso posterior.")
    return ";".join(to_list), ";".join(cc_list), should_send_at_end


def compose_validacao_email(periodo_label: str, arquivo_nome: str, executor: str) -> str:
    return (
        "<p>Prezados,</p>"
        f'<p>Segue arquivo <b>"{arquivo_nome}"</b> para validação do time Comercial.</p>'
        "<p><b>INSTRUÇÕES PARA VALIDAÇÃO:</b></p>"
        "<ul>"
        '  <li><b>ABA "ajuste":</b> Preencher APENAS esta aba com as correções necessárias'
        "    <ul>"
        "      <li>Atendimento_ajustado</li>"
        "      <li>Contrato_ajustado</li>"
        "      <li>Modalidade_ajustado</li>"
        "      <li>Inicio_ajustado</li>"
        "      <li>Final_ajustado</li>"
        "    </ul>"
        "  </li>"
        '  <li><b>IMPORTANTE:</b> Cada linha é única (combinação de Bacia Ajustada + Bloco Ajustado + Prefixo SAP)</li>'
        '  <li>Aba "cronograma_base": disponível para consulta (NÃO alterar)</li>'
        '  <li>Aba "Base Dados": disponível para consulta (NÃO alterar)</li>'
        '  <li>Aba "TD": tabela dinâmica para análise (NÃO alterar)</li>'
        "</ul>"
        '<p>Após validação e preenchimento da aba "ajuste", <b>retornar o arquivo para Guilherme</b>.</p>'
        "<p>Atenciosamente,<br>"
        f"{executor}</p>"
        f"<p>Período: {periodo_label}</p>"
    )


def ask_and_send_email(
    attachment_path: Path,
    periodo_label: str,
    logger: logging.Logger,
    to_value: str,
    cc_value: str,
) -> None:
    subject = f"Cronograma para Validação - {periodo_label}"
    body = compose_validacao_email(periodo_label, attachment_path.name, getpass.getuser())

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = to_value
    mail.CC = cc_value
    mail.Subject = subject
    mail.HTMLBody = body
    mail.Attachments.Add(str(attachment_path.resolve()))
    mail.Send()
    logger.info("E-mail de validação enviado para: %s", to_value)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera arquivo de validação do output_outlook.")
    parser.add_argument("--year", type=int, help="Ano para gerar (ex: 2026).")
    parser.add_argument("--month", type=int, help="Mês inicial para gerar (1-12).")
    parser.add_argument("--end-year", type=int, help="Ano final para gerar (ex: 2026). Padrão: ano corrente.")
    parser.add_argument("--end-month", type=int, help="Mês final para gerar (1-12). Padrão: 12 (dezembro).")
    parser.add_argument(
        "--selection-timeout",
        type=int,
        default=600,
        help="Tempo maximo (segundos) para selecao de ano/mes na interface (padrao: 600).",
    )
    parser.add_argument(
        "--to",
        type=str,
        help="Destinatários Para separados por ';' (pula coleta na interface).",
    )
    parser.add_argument(
        "--cc",
        type=str,
        default="",
        help="Destinatários Cc separados por ';' (pula coleta na interface).",
    )
    parser.add_argument("--no-email", action="store_true", help="Não envia e-mail ao final.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    logger = setup_logging()

    if not DE_PARA_FILE.exists():
        raise FileNotFoundError(f"Arquivo DE_PARA não encontrado: {DE_PARA_FILE}")

    if args.year and args.month:
        ano_base, mes_base = args.year, args.month
        ano_final = args.end_year if args.end_year else date.today().year
        mes_final = args.end_month if args.end_month else 12
        logger.info(
            "Periodo por parâmetro: inicio=%02d/%s | fim=%02d/%s",
            mes_base,
            ano_base,
            mes_final,
            ano_final,
        )
    else:
        periodo = pick_year_month(timeout_seconds=args.selection_timeout)
        if periodo is None:
            logger.warning("Execução cancelada na seleção de período.")
            return
        ano_base, mes_base, ano_final, mes_final = periodo
        logger.info(
            "Periodo selecionado pelo usuário: inicio=%02d/%s | fim=%02d/%s",
            mes_base,
            ano_base,
            mes_final,
            ano_final,
        )

    if mes_base < 1 or mes_base > 12:
        raise ValueError("Mês inválido. Use valor entre 1 e 12.")
    if mes_final < 1 or mes_final > 12:
        raise ValueError("Mês final inválido. Use valor entre 1 e 12.")

    start_ref = date(ano_base, mes_base, 1)
    end_ref = date(ano_final, mes_final, 1)
    if end_ref < start_ref:
        raise ValueError("Período final deve ser maior ou igual ao período inicial.")

    first_process_ref = next_month(start_ref)
    if end_ref < first_process_ref:
        raise ValueError(
            "Intervalo sem meses para processar. Ajuste o mês final para pelo menos o mês seguinte ao inicial."
        )

    email_recipients: tuple[str, str] | None = None
    should_send_email = False
    if not args.no_email:
        if args.to:
            to_list = _split_email_list(args.to)
            cc_list = _split_email_list(args.cc)
            all_recipients = to_list + cc_list
            if not to_list:
                raise ValueError("--to deve conter ao menos um destinatário.")
            valid, invalids = _validate_email_list(all_recipients)
            if not valid:
                raise ValueError(f"E-mails inválidos informados em parâmetros: {invalids}")
            email_recipients = (";".join(to_list), ";".join(cc_list))
            should_send_email = True
            logger.info("Destinatários recebidos por parâmetro e validados.")
        else:
            recipients_info = ask_and_confirm_recipients(logger)
            if recipients_info is None:
                logger.info("Execução cancelada antes do processamento por falta de confirmação dos e-mails.")
                return
            email_recipients = (recipients_info[0], recipients_info[1])
            should_send_email = recipients_info[2]

    periodo_label = f"{month_name_pt(mes_base)}/{ano_base} até {month_name_pt(mes_final)}/{ano_final}"

    base_dados_path = resolve_base_dados_file()
    logger.info("Base de dados: %s", base_dados_path)

    df_base = read_excel_with_fallback(base_dados_path, "Base Dados")
    df_de_para = read_excel_with_fallback(DE_PARA_FILE, "DE_PARA")

    de_para_map, nao_scoped_bacia_bloco = build_de_para_map(df_de_para, logger)
    df_cronograma = build_cronograma_base(
        df_base,
        de_para_map,
        nao_scoped_bacia_bloco,
        ano_base,
        mes_base,
        ano_final,
        mes_final,
        logger,
    )
    df_ajuste = build_ajuste_sheet(df_cronograma)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / "cronograma_outlook_validacao.xlsx"

    try:
        write_output_workbook(base_dados_path, df_cronograma, df_ajuste, output_path)
    except PermissionError:
        raise PermissionError(
            f"Nao foi possivel sobrescrever {output_path}. Feche o arquivo no Excel e execute novamente."
        )

    apply_cronograma_basic_format(output_path)
    apply_ajuste_format(output_path)
    try:
        build_td_pivot_with_slicers(output_path, logger)
    except Exception as exc:
        logger.warning("TD dinâmica não criada nesta execução. Seguindo com etapa 1. Motivo: %s", exc)

    input_copy_path = copy_output_to_input(output_path, logger)

    logger.info("Arquivo gerado: %s", output_path)
    logger.info("Arquivo de entrada do Script 3 atualizado: %s", input_copy_path)
    logger.info("Linhas no cronograma_base: %d", len(df_cronograma))

    if not args.no_email:
        if email_recipients is None:
            logger.info("Sem destinatários confirmados. Envio não realizado.")
            return
        if should_send_email:
            ask_and_send_email(output_path, periodo_label, logger, email_recipients[0], email_recipients[1])
        else:
            logger.info("Envio desativado pelo usuário no início do processo.")
    else:
        logger.info("Envio de e-mail desativado por --no-email.")


if __name__ == "__main__":
    main()
