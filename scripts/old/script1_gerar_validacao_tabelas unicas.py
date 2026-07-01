from __future__ import annotations

import argparse
import calendar
import getpass
import logging
import time
from dataclasses import dataclass
import re
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk

import pandas as pd
import win32com.client  # type: ignore[import]
from openpyxl.formatting.rule import FormulaRule
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parents[1]
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"

BASE_DADOS_FILE = INPUT_DIR / "Base_Dados.xlsx"
DE_PARA_FILE = INPUT_DIR / "DE_PARA.xlsx"

DEFAULT_TO = [
    "marcelo.knust.prestserv@transpetro.com.br",
    #"barcanias@transpetro.com.br",
    #"williamarthur@transpetro.com.br",
    #"mrnobrega@transpetro.com.br",
]
DEFAULT_CC = [
    "marcelo.knust.prestserv@transpetro.com.br",
    #"felipedan@transpetro.com.br",
]

CRONOGRAMA_COLS = [
    "TIPO",
    "Exclusões",
    "Bacia Ajustada",
    "Bloco Ajustado",
    "Prefixo SAP",
    "MesReferencia",
    "InicioNoMes",
    "FimNoMes",
    "DiasOperadosMes",
    "DiasOperadosMes_Rateado",
    "DiasIntervaloTotal",
    "Atendimento",
    "Contrato",
    "Modalidade",
    "Validado",
    "InicioNoMes_ajustado",
    "FimNoMes_ajustado",
    "Atendimento_ajustado",
    "Contrato_ajustado",
    "Modalidade_ajustado",
]


@dataclass(frozen=True)
class MonthSlice:
    mes_referencia: date
    inicio_no_mes: date
    fim_no_mes: date
    dias_operados_mes: int


def setup_logging() -> logging.Logger:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log_path = OUTPUT_DIR / f"script1_gerar_validacao_{datetime.now():%Y%m%d_%H%M%S}.log"

    logger = logging.getLogger("output_outlook.script1")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    logger.info("Log inicializado em: %s", log_path)
    return logger


def normalize_name(value: object) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = text.replace("á", "a").replace("à", "a").replace("â", "a").replace("ã", "a")
    text = text.replace("é", "e").replace("ê", "e")
    text = text.replace("í", "i")
    text = text.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    text = text.replace("ú", "u")
    text = text.replace("ç", "c")
    return " ".join(text.split())


def excel_value_to_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def first_day_of_month(value: date) -> date:
    return value.replace(day=1)


def last_day_of_month(value: date) -> date:
    return value.replace(day=calendar.monthrange(value.year, value.month)[1])


def next_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def previous_month(value: date) -> date:
    if value.month == 1:
        return date(value.year - 1, 12, 1)
    return date(value.year, value.month - 1, 1)


def month_ref_start_for_date(value: date) -> date:
    if value.day >= 26:
        return next_month(first_day_of_month(value))
    return first_day_of_month(value)


def build_month_period(month_ref_start: date) -> tuple[date, date, date]:
    prev_month = previous_month(month_ref_start)
    period_start = prev_month.replace(day=26)
    period_end = month_ref_start.replace(day=25)
    month_ref_end = last_day_of_month(month_ref_start)
    return period_start, period_end, month_ref_end


def split_interval_by_custom_month(start_date: date, end_date: date) -> list[MonthSlice]:
    slices: list[MonthSlice] = []
    cursor = month_ref_start_for_date(start_date)
    end_cursor = month_ref_start_for_date(end_date)

    while cursor <= end_cursor:
        period_start, period_end, month_ref_end = build_month_period(cursor)
        slice_start = max(start_date, period_start)
        slice_end = min(end_date, period_end)

        if slice_start <= slice_end:
            slices.append(
                MonthSlice(
                    mes_referencia=month_ref_end,
                    inicio_no_mes=slice_start,
                    fim_no_mes=slice_end,
                    dias_operados_mes=(slice_end - slice_start).days + 1,
                )
            )

        cursor = next_month(cursor)

    return slices


def pick_year_month(timeout_seconds: int = 600) -> tuple[int, int] | None:
    current = date.today().year
    years = [str(y) for y in range(current - 5, current + 6)]
    months = [
        ("01", "Janeiro"),
        ("02", "Fevereiro"),
        ("03", "Marco"),
        ("04", "Abril"),
        ("05", "Maio"),
        ("06", "Junho"),
        ("07", "Julho"),
        ("08", "Agosto"),
        ("09", "Setembro"),
        ("10", "Outubro"),
        ("11", "Novembro"),
        ("12", "Dezembro"),
    ]

    root = tk.Tk()
    root.title("Selecao de Periodo - output_outlook")
    root.geometry("460x250")
    root.resizable(False, False)
    root.attributes("-topmost", True)

    ttk.Label(root, text="Ano:").pack(anchor="w", padx=16, pady=(16, 4))
    selected_year = tk.StringVar(value=str(current))
    ttk.Combobox(root, values=years, textvariable=selected_year, state="readonly").pack(fill=tk.BOTH, padx=16)

    ttk.Label(root, text="Mes Inicial:").pack(anchor="w", padx=16, pady=(12, 4))
    month_display = [f"{num} - {name}" for num, name in months]
    selected_month = tk.StringVar(value=month_display[date.today().month - 1])
    ttk.Combobox(root, values=month_display, textvariable=selected_month, state="readonly").pack(fill=tk.BOTH, padx=16)

    result: dict[str, int] = {}
    remaining_seconds = max(30, int(timeout_seconds))
    timer_var = tk.StringVar()

    def format_remaining(seconds: int) -> str:
        mins = seconds // 60
        secs = seconds % 60
        return f"Tempo para selecao: {mins:02d}:{secs:02d}"

    def tick() -> None:
        nonlocal remaining_seconds
        timer_var.set(format_remaining(remaining_seconds))
        if remaining_seconds <= 0:
            messagebox.showwarning("Tempo esgotado", "Tempo para selecao encerrado. Execute novamente para escolher o periodo.")
            root.destroy()
            return
        remaining_seconds -= 1
        root.after(1000, tick)

    def confirm() -> None:
        try:
            result["year"] = int(selected_year.get())
            result["month"] = int(selected_month.get().split(" - ", 1)[0])
            root.destroy()
        except Exception:
            messagebox.showerror("Erro", "Selecione ano e mes validos.")

    def cancel() -> None:
        root.destroy()

    def extend_time() -> None:
        nonlocal remaining_seconds
        remaining_seconds += 300
        timer_var.set(format_remaining(remaining_seconds))

    buttons = ttk.Frame(root)
    buttons.pack(fill=tk.BOTH, padx=16, pady=(12, 10))
    ttk.Label(root, textvariable=timer_var).pack(anchor="w", padx=16, pady=(0, 4))
    ttk.Button(root, text="+5 min", command=extend_time).pack(anchor="w", padx=16)
    ttk.Button(buttons, text="Cancelar", command=cancel).pack(side=tk.RIGHT)
    ttk.Button(buttons, text="Confirmar", command=confirm).pack(side=tk.RIGHT, padx=(0, 8))

    tick()
    root.mainloop()

    if "year" not in result or "month" not in result:
        return None
    return result["year"], result["month"]


def resolve_base_dados_file() -> Path:
    if BASE_DADOS_FILE.exists():
        return BASE_DADOS_FILE

    candidates = sorted(INPUT_DIR.glob("*Base*Dados*.xlsx"))
    if candidates:
        return candidates[0]

    raise FileNotFoundError(
        f"Arquivo de base nao encontrado em {INPUT_DIR}. Esperado: {BASE_DADOS_FILE.name}"
    )


def read_excel_with_fallback(path: Path, preferred_sheet: str) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True)
    sheet = preferred_sheet if preferred_sheet in wb.sheetnames else wb.sheetnames[0]
    wb.close()
    return pd.read_excel(path, sheet_name=sheet)


def find_column(df: pd.DataFrame, aliases: list[str], required: bool = True) -> str | None:
    normalized = {normalize_name(c): str(c) for c in df.columns}
    for alias in aliases:
        col = normalized.get(normalize_name(alias))
        if col:
            return col

    if required:
        raise KeyError(f"Nao foi possivel localizar coluna. Tentativas: {aliases}")
    return None


def month_name_pt(month: int) -> str:
    names = [
        "Janeiro",
        "Fevereiro",
        "Marco",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    return names[month - 1]


def _clean_key_part(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _lookup_key_preferred(prefixo_unidade: object, bloco: object) -> tuple[str, str, str] | None:
    p = _clean_key_part(prefixo_unidade)
    b = _clean_key_part(bloco)
    if p and b:
        return ("prefixo", p, b)
    return None


def _lookup_key_fallback(bacia: object, bloco: object) -> tuple[str, str, str] | None:
    ba = _clean_key_part(bacia)
    bl = _clean_key_part(bloco)
    if ba and bl:
        return ("bacia", ba, bl)
    return None


def build_de_para_map(df_de_para: pd.DataFrame, logger: logging.Logger) -> dict[tuple[str, str, str], dict[str, object]]:
    col_bacia = find_column(df_de_para, ["Bacia Ajustado", "Bacia Ajustada"])
    col_bloco = find_column(df_de_para, ["Bloco Ajustado"])
    col_unidade = find_column(df_de_para, ["Unidade Marítima", "Unidade Maritima"], required=False)
    col_atendimento = find_column(df_de_para, ["Atendimento"], required=False)
    col_contrato = find_column(df_de_para, ["Contrato"], required=False)
    col_modalidade = find_column(df_de_para, ["Modalidade"], required=False)

    result: dict[tuple[str, str, str], dict[str, object]] = {}
    duplicates = 0
    for _, row in df_de_para.iterrows():
        unidade_val = row.get(col_unidade) if col_unidade else None
        bloco_val = row.get(col_bloco)
        bacia_val = row.get(col_bacia)

        key = _lookup_key_preferred(unidade_val, bloco_val)
        if key is None:
            key = _lookup_key_fallback(bacia_val, bloco_val)

        if key is None:
            continue
        if key in result:
            duplicates += 1
        result[key] = {
            "Atendimento": row.get(col_atendimento) if col_atendimento else None,
            "Contrato": row.get(col_contrato) if col_contrato else None,
            "Modalidade": row.get(col_modalidade) if col_modalidade else None,
        }

    if duplicates:
        logger.warning("DE_PARA contem %d chaves duplicadas (ultima ocorrencia prevalece).", duplicates)
    return result


def build_cronograma_base(
    df_base: pd.DataFrame,
    de_para_map: dict[tuple[str, str, str], dict[str, object]],
    ano_base: int,
    mes_base: int,
) -> pd.DataFrame:
    col_tipo = find_column(df_base, ["TIPO", "Tipo"])
    col_exclusoes = find_column(df_base, ["Exclusões", "Exclusoes", "Q"], required=False)
    col_bacia = find_column(df_base, ["Bacia Ajustada", "Bacia Ajustado"])
    col_bloco = find_column(df_base, ["Bloco Ajustado"])
    col_prefixo = find_column(df_base, ["Prefixo SAP", "Prefixo"])
    col_inicio = find_column(df_base, ["Data inicio", "Data Inicio", "Inicio", "Inicio Programado", "Data inicial"])
    col_fim = find_column(df_base, ["Data termino", "Data fim", "Fim", "Fim Programado", "Data final"])

    # AnoMesBase define o corte inicial; processa do mes seguinte ate dezembro do mesmo ano.
    first_ref_month_start = next_month(date(ano_base, mes_base, 1))
    min_ref = last_day_of_month(first_ref_month_start)
    max_ref = last_day_of_month(date(ano_base, 12, 1))

    rows_out: list[dict[str, object]] = []
    for _, row in df_base.iterrows():
        tipo_val = _clean_key_part(row.get(col_tipo)).upper()
        if tipo_val not in {"SONDA", "UEP", "UMS"}:
            continue

        start_date = excel_value_to_date(row.get(col_inicio))
        end_date = excel_value_to_date(row.get(col_fim))
        if start_date is None or end_date is None or end_date < start_date:
            continue

        month_slices = split_interval_by_custom_month(start_date, end_date)
        total_days = sum(item.dias_operados_mes for item in month_slices)

        for item in month_slices:
            if item.mes_referencia < min_ref or item.mes_referencia > max_ref:
                continue

            bacia_val = _clean_key_part(row.get(col_bacia, ""))
            bloco_val = _clean_key_part(row.get(col_bloco, ""))
            prefixo_val = _clean_key_part(row.get(col_prefixo, ""))

            preferred_key = _lookup_key_preferred(prefixo_val, bloco_val)
            fallback_key = _lookup_key_fallback(bacia_val, bloco_val)

            lookup_key = preferred_key if preferred_key is not None else fallback_key
            d = de_para_map.get(lookup_key, {}) if lookup_key else {}
            has_match = lookup_key in de_para_map if lookup_key else False
            atendimento_val = _clean_key_part(d.get("Atendimento")) or "(vazio)"
            contrato_val = _clean_key_part(d.get("Contrato")) or "(vazio)"
            modalidade_val = _clean_key_part(d.get("Modalidade")) or "(vazio)"

            rows_out.append(
                {
                    "TIPO": row.get(col_tipo),
                    "Exclusões": row.get(col_exclusoes) if col_exclusoes else None,
                    "Bacia Ajustada": bacia_val,
                    "Bloco Ajustado": bloco_val,
                    "Prefixo SAP": row.get(col_prefixo),
                    "MesReferencia": item.mes_referencia,
                    "InicioNoMes": item.inicio_no_mes,
                    "FimNoMes": item.fim_no_mes,
                    "DiasOperadosMes": item.dias_operados_mes,
                    "DiasOperadosMes_Rateado": float(item.dias_operados_mes),
                    "DiasIntervaloTotal": total_days,
                    "Atendimento": atendimento_val,
                    "Contrato": contrato_val,
                    "Modalidade": modalidade_val,
                    "Validado": "Sim" if not has_match else "Não",
                    "InicioNoMes_ajustado": None,
                    "FimNoMes_ajustado": None,
                    "Atendimento_ajustado": None,
                    "Contrato_ajustado": None,
                    "Modalidade_ajustado": None,
                }
            )

    if not rows_out:
        raise ValueError("Nenhuma linha encontrada para o período selecionado.")

    return pd.DataFrame(rows_out)[CRONOGRAMA_COLS]


def write_output_workbook(base_dados_path: Path, df_cronograma: pd.DataFrame, output_path: Path) -> None:
    wb = load_workbook(base_dados_path)
    base_sheet = wb["Base Dados"] if "Base Dados" in wb.sheetnames else wb[wb.sheetnames[0]]

    for sheet_name in list(wb.sheetnames):
        if wb[sheet_name] is not base_sheet:
            del wb[sheet_name]

    base_sheet.title = "Base Dados"

    ws_crono = wb.create_sheet("cronograma_base")
    ws_crono.append(CRONOGRAMA_COLS)
    for row in df_cronograma.itertuples(index=False, name=None):
        ws_crono.append(list(row))

    wb.create_sheet("TD")
    wb.save(output_path)
    wb.close()


def apply_cronograma_basic_format(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["cronograma_base"]

    max_row = ws.max_row
    max_col = ws.max_column

    table_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = table_ref

    header_map = {str(c.value): c.column for c in ws[1] if c.value is not None}
    for col_name in ["MesReferencia", "InicioNoMes", "FimNoMes", "InicioNoMes_ajustado", "FimNoMes_ajustado"]:
        col_idx = header_map.get(col_name)
        if col_idx:
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = "DD/MM/YYYY"

    for col_name in ["DiasOperadosMes", "DiasOperadosMes_Rateado", "DiasIntervaloTotal"]:
        col_idx = header_map.get(col_name)
        if col_idx:
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = "#,##0.00"

    validado_col = header_map.get("Validado")
    ajustado_cols = [
        header_map.get("InicioNoMes_ajustado"),
        header_map.get("FimNoMes_ajustado"),
        header_map.get("Atendimento_ajustado"),
        header_map.get("Contrato_ajustado"),
        header_map.get("Modalidade_ajustado"),
    ]

    if validado_col:
        validado_letter = ws.cell(row=1, column=validado_col).column_letter
        validado_range = f"{validado_letter}2:{validado_letter}{max_row}"

        dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=False)
        dv.error = "Use apenas Sim ou Não."
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)
        dv.add(validado_range)

        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=validado_col)
            if cell.value == "Não":
                cell.fill = green_fill
            elif cell.value == "Sim":
                cell.fill = red_fill

        ws.conditional_formatting.add(
            validado_range,
            FormulaRule(formula=[f'${validado_letter}2="Não"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            validado_range,
            FormulaRule(formula=[f'${validado_letter}2="Sim"'], fill=red_fill),
        )

    for col_idx in ajustado_cols:
        if not col_idx:
            continue
        for row_idx in range(2, max_row + 1):
            ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 34)

    wb.save(workbook_path)
    wb.close()


def build_td_pivot_with_slicers(workbook_path: Path, logger: logging.Logger) -> None:
    import os
    import subprocess
    import tempfile
    import zipfile

    # Estrategia de persistencia: gera pivot no arquivo aberto e salva uma
    # copia temporaria (equivale ao fluxo manual de salvar com novo nome).
    source_path = workbook_path.resolve()
    temp_save_path = source_path.with_name(f"{source_path.stem}__pivot_tmp{source_path.suffix}")
    file_path_escaped = str(source_path).replace("'", "''")
    saveas_path_escaped = str(temp_save_path).replace("'", "''")

    ps_script = f"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
    $wb = $excel.Workbooks.Open('{file_path_escaped}')
    $wsData = $wb.Worksheets('cronograma_base')
    $wsTD  = $wb.Worksheets('TD')
    $wsTD.Cells.Clear()

    # Remove caches de slicer preexistentes no arquivo para evitar conflito
    for ($ci = $wb.SlicerCaches.Count; $ci -ge 1; $ci--) {{
        try {{ $wb.SlicerCaches.Item($ci).Delete() }} catch {{}}
    }}

    $lastRow = $wsData.UsedRange.Rows.Count
    $lastCol = $wsData.UsedRange.Columns.Count
    $sourceRef = "cronograma_base!R1C1:R$($lastRow)C$($lastCol)"
    Write-Host "SOURCE: $sourceRef"
    $cache = $wb.PivotCaches().Create(1, $sourceRef)
    $pt = $cache.CreatePivotTable($wsTD.Range('A3'), 'PivotCronograma')
    Write-Host "CREATED: $($wsTD.PivotTables().Count)"
    $pt.PivotFields('Bloco Ajustado').Orientation = 1
    $pt.PivotFields('Prefixo SAP').Orientation     = 1
    $pfMes = $pt.PivotFields('MesReferencia')
    $pfMes.Orientation = 2
    [void]$pt.AddDataField($pt.PivotFields('DiasOperadosMes'), 'Soma DiasOperadosMes', -4157)
    # Forca formato local para evitar exibicao literal "yyyy" nos rotulos
    try {{ $pfMes.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    try {{ $pfMes.DataRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    try {{ $pfMes.LabelRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    $pt.RefreshTable() | Out-Null
    try {{ $pfMes.DataRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}
    try {{ $pfMes.LabelRange.NumberFormatLocal = 'mm/aaaa' }} catch {{}}

    $pivotRange = $pt.TableRange2
    $anchorLeft = $pivotRange.Left + $pivotRange.Width + 24
    $anchorTop = $pivotRange.Top
    $slicerWidth = 180
    $slicerHeight = 120
    $hGap = 10
    $vGap = 10
    $slicerFields = @('TIPO','Contrato','Modalidade','Atendimento')

    # Definicao clara da grade
    $columns = 2
    $rows = 2

    # Remove slicers antigos da aba TD para evitar sobreposicao entre execucoes
    for ($si = $wsTD.Shapes.Count; $si -ge 1; $si--) {{
        try {{
            $sh = $wsTD.Shapes.Item($si)
            if ($sh.Name -like 'Slicer_*' -or $sh.Name -like 'SL_*') {{
                $sh.Delete()
            }}
        }} catch {{}}
    }}

    # Posicionamento dos slicers
    for ($i = 0; $i -lt $slicerFields.Count; $i++) {{
        $field = $slicerFields[$i]
        try {{
            # Calcula linha e coluna (ordem: esquerda->direita, cima->baixo)
            $row = [math]::Floor($i / $columns)
            $col = $i % $columns

            # Cria o slicer
            $pfSlicer = $pt.PivotFields($field)
            $cacheName = ('SC_' + ($field -replace '[^A-Za-z0-9]', '') + '_' + $i)
            try {{
                $sc = $wb.SlicerCaches.Add2($pt, $pfSlicer, $cacheName)
            }} catch {{
                $sc = $wb.SlicerCaches.Add($pt, $pfSlicer, $cacheName)
            }}

            $sl = $sc.Slicers.Add($wsTD)

            # Posiciona na grade usando ancoras de celula (mais estavel no COM)
            $anchorCell = $wsTD.Cells((2 + ($row * 9)), (10 + ($col * 4)))
            $sl.Left = $anchorCell.Left
            $sl.Top = $anchorCell.Top
            $sl.Width = $slicerWidth
            $sl.Height = $slicerHeight

            Write-Host "Slicer '$field' posicionado em Linha $row, Coluna $col"

        }} catch {{
            Write-Warning "Erro ao criar slicer para '$field': $_"
        }}
    }}
    $pivotCount = $wsTD.PivotTables().Count
    Write-Host "PIVOT_BEFORE_CLOSE: $pivotCount"
    $wb.SaveCopyAs('{saveas_path_escaped}')
    if (-not (Test-Path '{saveas_path_escaped}')) {{
        throw 'SaveCopyAs nao gerou arquivo.'
    }}
    Write-Host "SAVED_PIVOT_COUNT: $pivotCount"
    $wb.Close($false)
}} finally {{
    $excel.Quit()
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""

    ps_file = Path(tempfile.mktemp(suffix=".ps1"))
    ps_file.write_text(ps_script, encoding="utf-8")
    try:
        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(ps_file)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
        stdout = (result.stdout or "").strip()
        stderr = (result.stderr or "").strip()
        logger.info("TD PowerShell stdout: %s", stdout)
        if stderr:
            logger.warning("TD PowerShell stderr: %s", stderr)
        if result.returncode != 0:
            raise RuntimeError(f"PowerShell falhou (rc={result.returncode}): {stderr}")
        if "SAVED_PIVOT_COUNT: 0" in stdout or "SAVED_PIVOT_COUNT:" not in stdout:
            raise RuntimeError(f"Pivot nao foi salva. Output: {stdout}")
        if not temp_save_path.exists():
            raise RuntimeError("PowerShell nao gerou arquivo temporario com pivot.")

        with zipfile.ZipFile(temp_save_path, "r") as zf:
            has_pivot_xml = any(name.startswith("xl/pivotTables/") for name in zf.namelist())
        if not has_pivot_xml:
            raise RuntimeError("Arquivo temporario salvo sem artefatos de pivot (xl/pivotTables).")

        os.replace(temp_save_path, source_path)
        logger.info("Aba TD criada como Tabela Dinâmica com filtros/slicers.")
    finally:
        if temp_save_path.exists():
            try:
                temp_save_path.unlink()
            except Exception:
                pass
        try:
            ps_file.unlink()
        except Exception:
            pass


def _split_email_list(raw_value: str) -> list[str]:
    return [item.strip() for item in raw_value.split(";") if item.strip()]


def _validate_email_list(recipients: list[str]) -> tuple[bool, list[str]]:
    pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    invalid = [email for email in recipients if not pattern.match(email)]
    return len(invalid) == 0, invalid


def ask_and_confirm_recipients(logger: logging.Logger) -> tuple[str, str, bool] | None:
    root = tk.Tk()
    root.withdraw()

    to_default = ";".join(DEFAULT_TO)
    cc_default = ";".join(DEFAULT_CC)

    to_value = simpledialog.askstring("Destino", "Destinatários (Para), separados por ';':", initialvalue=to_default)
    if to_value is None:
        root.destroy()
        logger.info("Coleta de e-mails cancelada pelo usuário.")
        return

    cc_value = simpledialog.askstring("Cópia", "Destinatários em cópia (Cc), separados por ';':", initialvalue=cc_default)
    if cc_value is None:
        root.destroy()
        logger.info("Coleta de e-mails cancelada pelo usuário.")
        return

    to_list = _split_email_list(to_value)
    cc_list = _split_email_list(cc_value)
    all_recipients = to_list + cc_list

    if not to_list:
        messagebox.showerror("E-mails inválidos", "Informe ao menos um destinatário em Para.")
        root.destroy()
        logger.warning("Lista de destinatários em Para vazia.")
        return None

    ok, invalids = _validate_email_list(all_recipients)
    if not ok:
        messagebox.showerror("E-mails inválidos", "Foram encontrados e-mails inválidos:\n" + "\n".join(invalids))
        root.destroy()
        logger.warning("E-mails inválidos informados: %s", invalids)
        return None

    confirm_msg = (
        "Confirma os destinatários abaixo para uso no envio final?\n\n"
        f"Para: {'; '.join(to_list)}\n"
        f"Cc: {'; '.join(cc_list) if cc_list else '(vazio)'}"
    )
    confirmed = messagebox.askyesno("Confirmar destinatários", confirm_msg)
    if not confirmed:
        root.destroy()
        logger.info("Destinatários não confirmados pelo usuário.")
        return None

    should_send_at_end = messagebox.askyesno(
        "Envio no Final",
        "Deseja enviar o e-mail no final do processo?",
    )

    root.destroy()
    logger.info("Destinatários validados e confirmados para uso posterior.")
    return ";".join(to_list), ";".join(cc_list), should_send_at_end


def compose_validacao_email(periodo_label: str, arquivo_nome: str, executor: str) -> str:
    return (
        "Prezados,\n\n"
        f"Segue arquivo \"{arquivo_nome}\" para validação do time Comercial.\n\n"
        "Por favor, validar as informações e preencher as colunas \"_ajustado\" conforme necessário.\n\n"
        "Após validação, retornar o arquivo para Guilherme.\n\n"
        "Atenciosamente,\n"
        f"{executor}\n"
        f"Período: {periodo_label}"
    )


def ask_and_send_email(
    attachment_path: Path,
    periodo_label: str,
    logger: logging.Logger,
    to_value: str,
    cc_value: str,
) -> None:
    subject = f"Cronograma para Validação - {periodo_label}"
    body = compose_validacao_email(periodo_label, attachment_path.name, getpass.getuser())

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = to_value
    mail.CC = cc_value
    mail.Subject = subject
    mail.Body = body
    mail.Attachments.Add(str(attachment_path.resolve()))
    mail.Send()
    logger.info("E-mail de validação enviado para: %s", to_value)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera arquivo de validação do output_outlook.")
    parser.add_argument("--year", type=int, help="Ano para gerar (ex: 2026).")
    parser.add_argument("--month", type=int, help="Mês inicial para gerar (1-12).")
    parser.add_argument(
        "--selection-timeout",
        type=int,
        default=600,
        help="Tempo maximo (segundos) para selecao de ano/mes na interface (padrao: 600).",
    )
    parser.add_argument(
        "--to",
        type=str,
        help="Destinatários Para separados por ';' (pula coleta na interface).",
    )
    parser.add_argument(
        "--cc",
        type=str,
        default="",
        help="Destinatários Cc separados por ';' (pula coleta na interface).",
    )
    parser.add_argument("--no-email", action="store_true", help="Não envia e-mail ao final.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    logger = setup_logging()

    if not DE_PARA_FILE.exists():
        raise FileNotFoundError(f"Arquivo DE_PARA não encontrado: {DE_PARA_FILE}")

    if args.year and args.month:
        ano_base, mes_base = args.year, args.month
        logger.info("AnoMesBase por parâmetro: ano=%s, mes=%s", ano_base, mes_base)
    else:
        periodo = pick_year_month(timeout_seconds=args.selection_timeout)
        if periodo is None:
            logger.warning("Execução cancelada na seleção de período.")
            return
        ano_base, mes_base = periodo
        logger.info("AnoMesBase selecionado pelo usuário: ano=%s, mes=%s", ano_base, mes_base)

    if mes_base < 1 or mes_base > 12:
        raise ValueError("Mês inválido. Use valor entre 1 e 12.")

    if mes_base == 12:
        raise ValueError("AnoMesBase em dezembro não possui meses seguintes no mesmo ano para processar.")

    email_recipients: tuple[str, str] | None = None
    should_send_email = False
    if not args.no_email:
        if args.to:
            to_list = _split_email_list(args.to)
            cc_list = _split_email_list(args.cc)
            all_recipients = to_list + cc_list
            if not to_list:
                raise ValueError("--to deve conter ao menos um destinatário.")
            valid, invalids = _validate_email_list(all_recipients)
            if not valid:
                raise ValueError(f"E-mails inválidos informados em parâmetros: {invalids}")
            email_recipients = (";".join(to_list), ";".join(cc_list))
            should_send_email = True
            logger.info("Destinatários recebidos por parâmetro e validados.")
        else:
            recipients_info = ask_and_confirm_recipients(logger)
            if recipients_info is None:
                logger.info("Execução cancelada antes do processamento por falta de confirmação dos e-mails.")
                return
            email_recipients = (recipients_info[0], recipients_info[1])
            should_send_email = recipients_info[2]

    periodo_label = f"{month_name_pt(mes_base)}/{ano_base} até Dezembro/{ano_base}"

    base_dados_path = resolve_base_dados_file()
    logger.info("Base de dados: %s", base_dados_path)

    df_base = read_excel_with_fallback(base_dados_path, "Base Dados")
    df_de_para = read_excel_with_fallback(DE_PARA_FILE, "DE_PARA")

    de_para_map = build_de_para_map(df_de_para, logger)
    df_cronograma = build_cronograma_base(df_base, de_para_map, ano_base, mes_base)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"cronograma_outlook_validacao_{ano_base}{mes_base:02d}.xlsx"

    try:
        write_output_workbook(base_dados_path, df_cronograma, output_path)
    except PermissionError:
        output_path = OUTPUT_DIR / f"cronograma_outlook_validacao_{ano_base}{mes_base:02d}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        logger.warning("Arquivo aberto. Usando fallback: %s", output_path)
        write_output_workbook(base_dados_path, df_cronograma, output_path)

    apply_cronograma_basic_format(output_path)
    try:
        build_td_pivot_with_slicers(output_path, logger)
    except Exception as exc:
        logger.warning("TD dinâmica não criada nesta execução. Seguindo com etapa 1. Motivo: %s", exc)

    logger.info("Arquivo gerado: %s", output_path)
    logger.info("Linhas no cronograma_base: %d", len(df_cronograma))

    if not args.no_email:
        if email_recipients is None:
            logger.info("Sem destinatários confirmados. Envio não realizado.")
            return
        if should_send_email:
            ask_and_send_email(output_path, periodo_label, logger, email_recipients[0], email_recipients[1])
        else:
            logger.info("Envio desativado pelo usuário no início do processo.")
    else:
        logger.info("Envio de e-mail desativado por --no-email.")


if __name__ == "__main__":
    main()
